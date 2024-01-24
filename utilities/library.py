import json

import openpyxl


class File_Operations:
    def __init__(self):
        global sheet
        global users_data_file
        work_book = openpyxl.load_workbook("/usr/local/google/home/sateeshg/apiautomation/apitestdata.xlsx")
        sheet = work_book["Sheet1"]
        # users_data_file = open("/usr/local/google/home/sateeshg/createuser", "r")

    def getRowCount(self):
        row = sheet.max_row
        print("Row Count: ", row)
        return row

    def getColumnCount(self):
        coln = sheet.max_column
        print("Column Count: ", coln)
        return coln

    def getKeyNamesList(self):
        keynames_li = []
        for c in range(1, (File_Operations.getColumnCount(self) + 1)):
            key_name = sheet.cell(row=1, column=c)
            # print("Key Names is: ", key_name.value)
            keynames_li.append(key_name.value)
        print("Key Names List Is: ", keynames_li)
        return keynames_li

    def getJsonData(self, users_data_file):
        users_jsondata = json.loads(users_data_file.read())
        print("Users Data from file in json format: ", users_jsondata)
        # js_name = users_jsondata['name']
        # print("Name Is: ", js_name)
        return users_jsondata

    def updateRequestWithData(self, r_no, users_jsondata, keynames_li):
        # c = 1
        for c in range(1, (File_Operations.getColumnCount(self) + 1)):
            cell_val = sheet.cell(row=r_no, column=c)

            users_jsondata[File_Operations().getKeyNamesList()[c - 1]] = cell_val.value
            print("Test: ", users_jsondata[keynames_li[c - 1]])
        return users_jsondata


# eo = File_Operations()
# eo.getJsonData(open("/usr/local/google/home/sateeshg/createuser", "r"))
# eo.getColumnCount()
# eo.getRowCount()
# eo.getKeyNamesList()
# eo.updateRequestWithData(4, eo.getJsonData(), eo.getKeyNamesList())
